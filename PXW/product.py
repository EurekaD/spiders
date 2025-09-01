import requests
from bs4 import BeautifulSoup
import re
import io
import pandas as pd
from glob import glob
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as DrawingImage
from PIL import Image as PILImage
import os


result_path = "结果.xlsx"


def append_to_excel(index, sku, price, img_path):
    """
    用 pandas 写入数据，并用 openpyxl 插入图片。
    """
    # 1. 先处理表格数据
    if os.path.exists(result_path):
        df = pd.read_excel(result_path)
    else:
        df = pd.DataFrame(columns=["Index", "SKU", "Price", "Image"])

    # 追加新行
    df = pd.concat([df, pd.DataFrame([{
        "Index": index,
        "SKU": sku,
        "Price": price,
        "Image": img_path
    }])], ignore_index=True)

    # 保存 Excel（先保存数据）
    df.to_excel(result_path, index=False)


def build_css_maps(soup):
    """从 <style> 标签中提取:
    1. 变量定义 (--xxx: url(...))
    2. class 映射 (.className { background-image:url(...) })
    """
    css_text = " ".join([s.get_text() for s in soup.find_all("style")])
    var_map = {}
    class_map = {}

    # 匹配变量定义
    for m in re.finditer(r"(--[a-zA-Z0-9_-]+)\s*:\s*url\(([^)]+)\)", css_text):
        var_map[m.group(1)] = m.group(2).strip('"\'')

    # 匹配 class 背景图
    for m in re.finditer(r"\.([a-zA-Z0-9_-]+)[^{]+{[^}]*background-image:\s*url\(([^)]+)\)", css_text):
        class_map[m.group(1)] = m.group(2).strip('"\'')

    return var_map, class_map


def extract_image(img, soup, var_map=None, class_map=None):
    """综合提取一个 <img> 的图片 URL"""
    # 初始化 CSS 映射（只构建一次）
    if var_map is None or class_map is None:
        var_map, class_map = build_css_maps(soup)

    # 情况1: 直接 src
    src = img.get("src")
    if src and not src.startswith("data:image/svg+xml"):
        return src.strip()

    # 情况2: style 内 background-image
    style = img.get("style", "")
    m = re.search(r"background-image:\s*url\(([^)]+)\)", style)
    if m:
        return m.group(1).strip('"\'')

    # 情况3: style 内 var(--xxx)
    m = re.search(r"background-image:\s*var\((--[^)]+)\)", style)
    if m:
        var_name = m.group(1)
        return var_map.get(var_name)

    # 情况4: class 对应背景图
    for c in img.get("class", []):
        if c in class_map:
            return class_map[c]

    return None


def extract_sku_from_href(href: str) -> str | None:
    """
    从 href 中提取 slug 和 SKU。
    示例 input: '/product/lestnitsa-stremyanka-stalnaya-4-stupeni-2199789552/?at=xxx'
    返回 SKU: '2199789552'，若未找到则返回 None。
    """
    try:
        # 提取 /product/ 和 /? 之间的部分
        m = re.search(r"/product/([^/]+)/?", href)
        if not m:
            return None

        slug = m.group(1)           # 整个 slug
        sku = slug.split("-")[-1]   # 取最后一部分
        return sku
    except Exception:
        return None


def get_html_file():
    """
    查找当前目录下的 HTML 文件。
    如果有，就返回第一个找到的文件路径。
    否则要求用户输入文件路径，并检查是否存在且是 HTML 文件。
    """
    # 查找当前目录下所有 HTML 文件
    html_files = glob("*.html")
    if html_files:
        print(f"找到 HTML 文件: {html_files[0]}")
        return html_files[0]

    # 如果没找到，要求用户输入路径
    while True:
        path = input("未找到 HTML 文件，请输入 HTML 文件完整路径: ").strip()
        if not os.path.exists(path):
            print("路径不存在，请重新输入。")
            continue
        if not path.lower().endswith(".html"):
            print("文件不是 HTML 文件，请重新输入。")
            continue
        return path


def col_width_to_pixels(width: float) -> int:
    """列宽转像素"""
    return int(width * 7) if width else 64


def row_height_to_pixels(height: float) -> int:
    """行高转像素"""
    return int(height * 1.33) if height else 20


def insert_images_from_column(file_path: str, sheet_name: str, image_col: str = "D", start_row: int = 2,
                              target_col_width: float = 25, target_row_height: float = 120):
    """
    Excel 指定列填充图片（覆盖原值 + 自动调整单元格）
    :param file_path: Excel 文件路径
    :param sheet_name: Sheet 名称
    :param image_col: 图片路径所在列
    :param start_row: 数据起始行（跳过表头）
    :param target_col_width: 插入图片的列宽（Excel 单位）
    :param target_row_height: 插入图片的行高（Excel 单位）
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 设置该列统一列宽
    ws.column_dimensions[image_col].width = target_col_width

    for row in range(start_row, ws.max_row + 1):
        cell = ws[f"{image_col}{row}"]
        img_path = cell.value

        if not img_path or not os.path.exists(img_path):
            continue

        # 删除原值
        cell.value = None

        # 设置该行行高
        ws.row_dimensions[row].height = target_row_height

        # 获取单元格可用像素大小
        col_pixels = col_width_to_pixels(target_col_width)
        row_pixels = row_height_to_pixels(target_row_height)

        # 获取图片原始大小
        with PILImage.open(img_path) as pil_img:
            img_w, img_h = pil_img.size

        # 按比例缩放，最大化填充单元格
        scale = min(col_pixels / img_w, row_pixels / img_h)
        new_w, new_h = int(img_w * scale), int(img_h * scale)

        # 插入图片
        img = DrawingImage(img_path)
        img.width, img.height = new_w, new_h
        ws.add_image(img, f"{image_col}{row}")

    # 保存文件
    out_path = file_path.replace(".xlsx", ".xlsx")
    wb.save(out_path)
    print(f"✅ 图片已插入完成，输出文件: {out_path}")


if __name__ == "__main__":

    path = get_html_file()

    # 读取本地 HTML 文件
    with open(path, "r", encoding="utf-8") as f:
        html = f.read()

    # 解析 HTML
    soup = BeautifulSoup(html, "html.parser")

    # 找到 paginator
    paginator = soup.find("div", id="paginator")

    # 找到所有 tileGridDesktop
    tile_grids = paginator.find_all("div", attrs={"data-widget": "tileGridDesktop"})

    # 输出目录
    os.makedirs("images", exist_ok=True)

    # 解析 css
    var_map, class_map = build_css_maps(soup)

    index = 1

    for grid in tile_grids:
        # 找到 grid 下的所有商品卡片（假设就是 div）
        product_cards = grid.find_all("div", recursive=False)

        for card in product_cards:

            sku = ""
            price = ""
            file_path = ""

            # Step 1 图片
            try:
                img = card.find("img")  # 每个商品卡片唯一的 img
                if not img:
                    continue

                url = extract_image(img, soup, var_map, class_map)
                if not url:
                    continue

                # 下载保存
                ext = os.path.splitext(url.split("?")[0])[1] or ".jpg"

                if url.startswith("http"):  # 网络地址
                    try:
                        resp = requests.get(url, timeout=10)
                        data = resp.content
                    except Exception as e:
                        print(f"下载图片失败 {url}: {e}")
                        continue

                elif url.startswith("data:image/"):  # base64
                    try:
                        import base64
                        data = base64.b64decode(url.split(",")[-1])
                    except Exception as e:
                        print(f"解析 base64 图片失败: {e}")
                        continue

                # 保存为 PNG
                try:
                    image = PILImage.open(io.BytesIO(data))
                    file_path = f"images/{index}.png"
                    image.save(file_path, format="PNG")
                    print(f"图片保存 {file_path}")
                except Exception as e:
                    print(f"保存图片失败: {e}")
                    continue

            except Exception as e:
                print(f"处理图片出错: {e}")
                continue

            # Step 2 href  sku
            try:
                a_tag = card.find("a", href=True)
                if a_tag:
                    href = a_tag["href"]
                    # print(href)
                    sku = extract_sku_from_href(href)
                    print(f"SKU {sku}")
            except Exception as e:
                print(f"提取 SKU 失败: {e}")

            # Step 3 价格
            try:
                price_tag = card.select_one("span.tsHeadline500Medium")
                if price_tag:
                    price = price_tag.get_text(strip=True)
                    print(f"价格 {price}")
            except Exception as e:
                print(f"提取价格失败: {e}")

            # Step 4 写入 Excel
            try:
                append_to_excel(index, sku, price, img_path=file_path)
            except Exception as e:
                print(f"写入 Excel 失败: {e}")

            print()
            index += 1

    print(f"总计解析 {index} 条数据")
    insert_images_from_column(result_path, sheet_name="Sheet1", image_col="D", start_row=2)
