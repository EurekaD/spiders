import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as DrawingImage
from PIL import Image as PILImage


def insert_images_from_column(file_path: str, image_col: str, start_row: int = 2,
                              target_col_width: float = 25, target_row_height: float = 120):
    """
    Excel 指定列填充图片（覆盖原值 + 自动调整单元格）

    :param file_path: Excel 文件路径
    :param image_col: 图片路径所在列
    :param start_row: 数据起始行（跳过表头）
    :param target_col_width: 插入图片的列宽（Excel 单位）
    :param target_row_height: 插入图片的行高（Excel 单位）
    """

    # 内部工具函数：列宽转像素
    def col_width_to_pixels(width: float) -> int:
        return int(width * 7) if width else 64

    # 内部工具函数：行高转像素
    def row_height_to_pixels(height: float) -> int:
        return int(height * 1.33) if height else 20

    wb = load_workbook(file_path)
    ws = wb.active

    # 设置该列统一列宽
    ws.column_dimensions[image_col].width = target_col_width

    for row in range(start_row, ws.max_row + 1):
        cell = ws[f"{image_col}{row}"]
        image_path = cell.value

        if not image_path or not os.path.exists(image_path):
            continue

        # 删除原值
        cell.value = None

        # 设置该行行高
        ws.row_dimensions[row].height = target_row_height

        # 获取单元格可用像素大小
        col_pixels = col_width_to_pixels(target_col_width)
        row_pixels = row_height_to_pixels(target_row_height)

        # 获取图片原始大小
        with PILImage.open(image_path) as pil_img:
            img_w, img_h = pil_img.size

        # 按比例缩放，最大化填充单元格
        scale = min(col_pixels / img_w, row_pixels / img_h)
        new_w, new_h = int(img_w * scale), int(img_h * scale)

        # 插入图片
        img = DrawingImage(image_path)
        img.width, img.height = new_w, new_h
        ws.add_image(img, f"{image_col}{row}")

    # 保存文件（避免覆盖原始文件）
    out_path = file_path.replace(".xlsx", "_with_images.xlsx")
    wb.save(out_path)
    print(f"图片已插入完成，输出文件: {out_path}")
