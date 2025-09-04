import time
from DrissionPage import ChromiumPage
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import os, sys
import mimetypes
from PIL import Image as PilImage
import random
import pandas as pd
import re
from openpyxl.utils import get_column_letter
import io
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# 手动添加 webp 支持
mimetypes.init()
mimetypes.types_map['.webp'] = 'image/webp'

# 创建全局锁，用于保护Excel写入操作
excel_lock = threading.Lock()
# 创建浏览器锁，用于安全地创建新标签页
browser_lock = threading.Lock()


def compress_image(img_path, quality=85, max_size=(800, 800)):
    """
    压缩图片并返回压缩后的图片数据

    参数:
    - img_path: 图片文件路径
    - quality: 压缩质量 (0-100)
    - max_size: 最大尺寸 (宽, 高)

    返回:
    - 压缩后的图片二进制数据
    """
    try:
        with PilImage.open(img_path) as img:
            # 转换模式（如果需要）
            if img.mode == 'RGBA':
                img = img.convert('RGB')

            # 调整尺寸
            img.thumbnail(max_size, PilImage.LANCZOS)

            # 创建内存缓冲区
            img_byte_arr = io.BytesIO()

            # 保存为JPEG格式（压缩率高）
            img.save(img_byte_arr, format='JPEG', quality=quality, optimize=True)

            return img_byte_arr.getvalue()

    except Exception as e:
        print(f"图片压缩失败: {e}")
        return None


def save_to_excel(img_paths, cell_positions, run_key_path, output_filename, product_url, row_number,
                  status='', run_key_data=None, run_key_titles=None,
                  green_price='', gray_price='', shop_description='', price_html=''):
    """
    将产品信息和图片保存到Excel文件

    参数:
    - img_paths: 图片路径列表
    - cell_positions: 图片要插入的单元格位置列表
    - output_filename: 输出的Excel文件名
    - product_url: 产品URL
    - row_number: 写入的行号
    - status: 商品状态（如售罄信息）
    - run_key_data: run_key表对应行的数据
    - run_key_titles: run_key表的列标题
    - green_price: 绿色价格
    - gray_price: 灰色价格
    - shop_description: 店铺介绍
    - price_html: 价格区域的HTML内容
    """
    # 使用全局锁保护Excel操作
    with excel_lock:
        # 如果文件不存在，创建带标题的Excel
        if not os.path.exists(output_filename):
            wb = Workbook()

            # 创建"原始数据表"sheet
            ws_original = wb.active
            ws_original.title = "原始数据表"

            # 创建"处理后数据"sheet
            ws_processed = wb.create_sheet("处理后数据")

            # 读取run_key原始数据
            if os.path.exists(run_key_path):
                try:
                    # 读取整个run_key表（不跳过任何行）
                    df_run_key = pd.read_excel(run_key_path, header=None)
                    print(f"成功读取run_key文件，共找到 {len(df_run_key)} 行数据")

                    # 将run_key数据写入"原始数据表"sheet
                    for r_idx, row in df_run_key.iterrows():
                        for c_idx, value in enumerate(row):
                            ws_original.cell(row=r_idx + 1, column=c_idx + 1, value=value)

                    print(f"run_key数据已写入原始数据表")

                    # 获取run_key标题（假设第6行是标题）
                    if len(df_run_key) >= 6:
                        run_key_titles = df_run_key.iloc[5].tolist()
                        print(f"获取run_key标题行: {run_key_titles}")
                    else:
                        run_key_titles = None
                        print("警告：无法获取run_key标题行")

                except Exception as e:
                    print(f"读取run_key文件时出错: {e}")
                    run_key_titles = None
            else:
                print(f"错误：找不到产品链接文件 {run_key_path}")
                run_key_titles = None

            # 在"处理后数据"sheet创建标题行
            if run_key_titles:
                # A-G列标题
                for col_idx in range(7):
                    col_letter = get_column_letter(col_idx + 1)
                    title = run_key_titles[col_idx] if col_idx < len(run_key_titles) else f"列{col_idx + 1}"
                    ws_processed[f'{col_letter}1'] = title

                # H列标题
                ws_processed['H1'] = "sku"

                # I-K列标题
                ws_processed['I1'] = "图片1"
                ws_processed['J1'] = "图片2"
                ws_processed['K1'] = "图片3"

                # L列标题 - 绿色价格
                ws_processed['L1'] = "绿色价格"

                # M列标题 - 灰色价格
                ws_processed['M1'] = "灰色价格"

                # N列标题 - 店铺介绍
                ws_processed['N1'] = "店铺介绍"

                # O列标题
                ws_processed['O1'] = "商品状态"

                # P-AO列标题 (run_key表的H-AG列)
                # H-AG列索引: 7-32 (共26列)
                # P列索引: 16, AO列索引: 41 (16+25=41)
                for src_idx in range(7, 33):  # H-AG列
                    if src_idx < len(run_key_titles):
                        target_col_idx = 16 + (src_idx - 7)  # 从P列(16)开始
                        target_col_letter = get_column_letter(target_col_idx)
                        title = run_key_titles[src_idx]
                        ws_processed[f'{target_col_letter}1'] = title
            else:
                # 如果没有标题，使用默认标题
                ws_processed['A1'] = "A列"
                ws_processed['B1'] = "B列"
                ws_processed['C1'] = "C列"
                ws_processed['D1'] = "D列"
                ws_processed['E1'] = "E列"
                ws_processed['F1'] = "F列"
                ws_processed['G1'] = "G列"
                ws_processed['H1'] = "sku"
                ws_processed['I1'] = "图片1"
                ws_processed['J1'] = "图片2"
                ws_processed['K1'] = "图片3"
                ws_processed['L1'] = "绿色价格"  # 新增
                ws_processed['M1'] = "灰色价格"  # 新增
                ws_processed['N1'] = "店铺介绍"  # 新增
                ws_processed['O1'] = "商品状态"

                # 添加P-AO列的默认标题
                for col_idx in range(16, 42):  # P(16)到AO(41)
                    col_letter = get_column_letter(col_idx)
                    ws_processed[f'{col_letter}1'] = f"列{col_letter}"

            # 添加AP列标题 - 价格HTML
            ws_processed['AP1'] = "价格HTML"

            wb.save(output_filename)
            print(f"初始化Excel文件：{output_filename}")
            # 注意：这里没有return语句，继续执行下面的代码

        temp_img_files = []  # 存储临时图片文件路径

        try:
            wb = load_workbook(output_filename)
            ws_processed = wb["处理后数据"]

            # 从URL提取sku
            sku = re.search(r'/(\d+)$', product_url)
            sku_value = sku.group(1) if sku else ""

            # 写入run_key数据到A-G列
            if run_key_data is not None and len(run_key_data) >= 7:
                for col_idx in range(7):  # A-G列
                    col_letter = get_column_letter(col_idx + 1)  # A=1, B=2, ..., G=7
                    cell = f"{col_letter}{row_number}"
                    ws_processed[cell] = run_key_data[col_idx]

            # 写入sku到H列
            ws_processed[f'H{row_number}'] = sku_value
            print(f"SKU已写入 H{row_number}: {sku_value}")

            # 写入绿色价格到L列
            if green_price:
                # 去除所有非数字字符
                green_price_clean = re.sub(r'[^\d]', '', green_price)
                ws_processed[f'L{row_number}'] = green_price_clean
                print(f"绿色价格已写入 L{row_number}: {green_price_clean}")
            else:
                print("无绿色价格信息")

            # 写入灰色价格到M列
            if gray_price:
                # 去除所有非数字字符
                gray_price_clean = re.sub(r'[^\d]', '', gray_price)
                ws_processed[f'M{row_number}'] = gray_price_clean
                print(f"灰色价格已写入 M{row_number}: {gray_price_clean}")
            else:
                print("无灰色价格信息")

            # 写入店铺介绍到N列
            if shop_description:
                ws_processed[f'N{row_number}'] = shop_description
                print(f"店铺介绍已写入 N{row_number}: {shop_description[:30]}...")  # 只显示前30个字符
            else:
                print("无店铺介绍信息")

            # 写入商品状态到O列
            ws_processed[f'O{row_number}'] = status
            print(f"商品状态已写入 O{row_number}: {status}")

            # 写入run_key数据到P-AO列 (H-AG列)
            if run_key_data is not None and len(run_key_data) >= 33:
                for col_idx in range(7, 33):  # H-AG列
                    # 目标列: P(16) + (col_idx - 7)
                    target_col_idx = 16 + (col_idx - 7)  # 16对应P列
                    target_col_letter = get_column_letter(target_col_idx)
                    cell = f"{target_col_letter}{row_number}"
                    ws_processed[cell] = run_key_data[col_idx]

            # 写入价格HTML到AP列
            if price_html:
                # 截断过长的HTML内容
                if len(price_html) > 32767:  # Excel单元格最大字符限制
                    price_html = price_html[:32700] + "...[截断]"
                ws_processed[f'AP{row_number}'] = price_html
                print(f"价格HTML已写入 AP{row_number}")
            else:
                print("无价格HTML信息")

            # 只处理有图片的情况
            for img_path, cell_position in zip(img_paths, cell_positions):
                if not os.path.exists(img_path):
                    print(f"警告：尝试处理的图片文件不存在：{img_path}")
                    continue

                try:
                    # 压缩图片
                    compressed_img_data = compress_image(img_path)

                    if compressed_img_data is None:
                        print(f"图片压缩失败，跳过: {img_path}")
                        continue

                    # 创建内存中的图片对象
                    img_byte_arr = io.BytesIO(compressed_img_data)

                    # 创建唯一的临时文件名
                    temp_img_path = os.path.join(os.path.dirname(img_path),
                                                 f"compressed_{os.path.basename(img_path)}_{random.randint(1000, 9999)}.jpg")

                    # 将压缩后的图片保存到临时文件
                    with open(temp_img_path, 'wb') as f:
                        f.write(compressed_img_data)

                    temp_img_files.append(temp_img_path)
                    print(f"图片已成功压缩并保存为 {temp_img_path}")

                    # 加载图片对象
                    try:
                        img_obj = Image(temp_img_path)
                        # 将图片插入到指定单元格
                        ws_processed.add_image(img_obj, cell_position)
                        print(f"图片已成功插入到 {output_filename} 的 {cell_position} 单元格")

                    except Exception as e:
                        print(f"加载图片失败: {e}")
                        continue

                    # 短暂延迟，确保文件操作完成
                    time.sleep(0.2)

                except Exception as e:
                    print(f"处理图片时发生错误：{e}")
                    continue

            # 保存工作簿
            wb.save(output_filename)
            print(f"Excel文件已保存: {output_filename}")

        except Exception as e:
            print(f"处理图片时发生错误：{e}")
        finally:
            # 关闭工作簿（如果已打开）
            if 'wb' in locals():
                try:
                    wb.close()
                    print("工作簿已关闭")
                except:
                    pass

            # 延迟一段时间确保文件释放
            time.sleep(0.5)

            # 删除临时图片文件
            for img_file in temp_img_files:
                try:
                    if os.path.exists(img_file):
                        os.remove(img_file)
                        print(f"已删除临时图片文件：{img_file}")
                    else:
                        print(f"临时图片文件未找到：{img_file}")
                except Exception as e:
                    print(f"删除临时图片文件时发生错误：{e}")

def process_single_url(url, idx, output_filename, image_dir, run_key_row_data,
                       run_key_titles, run_key_path, max_workers, main_page, excel_row):
    """
    处理单个URL的函数，用于多线程执行

    参数:
    - url: 要处理的URL
    - idx: URL的索引号
    - output_filename: 输出Excel文件名
    - image_dir: 图片保存目录
    - run_key_row_data: run_key表对应行的数据
    - run_key_titles: run_key表的列标题
    - run_key_path: run_key文件路径
    - max_workers: 最大线程数
    - main_page: 主浏览器对象
    - excel_row: 在原始Excel中的行号
    """
    print(f"\n开始处理链接 [{idx + 1}]: {url}")

    # 跳过空链接
    if pd.isna(url) or not str(url).startswith('http'):
        print(f"跳过无效链接: {url}")
        return

    status = ""  # 初始化商品状态
    img_paths = []  # 初始化图片路径列表
    green_price = ""  # 初始化绿色价格
    gray_price = ""  # 初始化灰色价格
    shop_description = ""  # 初始化店铺介绍
    price_html = ""  # 初始化价格HTML

    try:
        # 使用浏览器锁安全地创建新标签页
        with browser_lock:
            tab = main_page.new_tab(url=url)
            print(f"[{idx + 1}] 已创建新标签页，处理URL: {url}")

        # 等待页面加载
        tab.wait.load_start()
        time.sleep(1)

        # 检查商品是否售罄
        if "Этот товар закончился" in tab.html:
            print(f"[{idx + 1}] 商品已售罄，记录状态...")
            status = "Этот товар закончился"  # 设置售罄状态

            # 保存商品信息（无图片）
            save_to_excel(
                img_paths=[],
                cell_positions=[],
                output_filename=output_filename,
                product_url=url,
                row_number=idx + 2,
                status=status,
                run_key_data=run_key_row_data,
                run_key_titles=run_key_titles,
                green_price=green_price,
                gray_price=gray_price,
                shop_description=shop_description,
                price_html=price_html,
                run_key_path=run_key_path
            )
            # 关闭标签页
            tab.close()
            return

        img_paths = []
        cell_positions = []
        i = 1  # 图片计数器

        # 定位所有图片容器
        img_containers = tab.eles('xpath://div[@class="pdp_ar5"]/div')
        print(f"[{idx + 1}] 找到 {len(img_containers)} 个图片容器")

        for container in img_containers:
            if i > 3:  # 只取前3张图片
                break

            if "svg" in str(container.inner_html).strip() or str(container.inner_html).strip() == "":
                continue

            try:
                img_element = container.ele('tag:img', timeout=5)
                if not img_element:
                    continue
                img_element.click()
                time.sleep(0.5)

                # 从URL提取sku
                sku = re.search(r'/(\d+)', url)
                sku_value = sku.group(1) if sku else f"unknown_{idx}"

                # 使用SKU命名图片
                img_name = f"{sku_value}_{i}.png"
                img_path = os.path.join(image_dir, img_name)

                # 保存图片  //div/img[contains(@alt, 'Лестница-стремянка 5 ступеней с широкими антискользящими ступенями')]
                tab.ele('xpath://div[@class="pdp_v3 pdp_v4"]/img').save(path=image_dir, name=img_name)
                print(f"[{idx + 1}] 图片保存到: {img_path}")

                # 添加到处理列表 - 单元格位置为：I、J、K列（同一行）
                img_paths.append(img_path)

                # 确定单元格位置：第一个产品在行2，第二个在行3，以此类推
                row_num = idx + 2  # 第一行是标题，所以从第二行开始
                # I=9, J=10, K=11
                column_letter = get_column_letter(8 + i)  # I=9, J=10, K=11
                cell_position = f"{column_letter}{row_num}"
                cell_positions.append(cell_position)
                i += 1
                time.sleep(0.5)  # 短暂等待

            except Exception as e:
                print(f"[{idx + 1}] 处理图片容器时出错: {e}")
                continue

        # 获取绿色价格
        try:
            green_price_ele = tab.ele('xpath://span[contains(text(), "c Ozon Картой")]/preceding-sibling::div/span', timeout=5)
            if green_price_ele:
                green_price = green_price_ele.text
                print(f"[{idx + 1}] 绿色价格: {green_price}")
            else:
                print(f"[{idx + 1}] 未找到绿色价格元素")
        except Exception as e:
            print(f"[{idx + 1}] 获取绿色价格时出错: {e}")

        # 获取灰色价格
        try:
            gray_price_ele = tab.ele('xpath://div[span[contains(text(), "без Ozon Карты")]]/preceding-sibling::div/span[1]', timeout=5)
            if gray_price_ele:
                gray_price = gray_price_ele.text
                print(f"[{idx + 1}] 灰色价格: {gray_price}")
            else:
                print(f"[{idx + 1}] 未找到灰色价格元素")
        except Exception as e:
            print(f"[{idx + 1}] 获取灰色价格时出错: {e}")

        # 获取价格区域HTML
        try:
            price_container = tab.ele('xpath://div[@class="pdp_h9b"]', timeout=5)
            if price_container:
                price_html = price_container.html
                print(f"[{idx + 1}] 成功获取价格区域HTML")
                # 打印部分HTML内容用于调试
                print(f"[{idx + 1}] 价格HTML片段: {price_html[:100]}...")
            else:
                print(f"[{idx + 1}] 未找到价格区域元素")
        except Exception as e:
            print(f"[{idx + 1}] 获取价格HTML时出错: {e}")

        # 点击店铺按钮获取店铺介绍
        try:
            shop_button = tab.ele('xpath://div[contains(text(), "Перейти в магазин")]/preceding-sibling::div//button', timeout=5)
            if shop_button:
                shop_button.click()
                time.sleep(0.5)

                # 获取店铺介绍
                shop_desc_ele = tab.ele('xpath://div[@class="ea5_3_1-a2 ea5_3_1-b2"]', timeout=5)
                if shop_desc_ele:
                    shop_description = shop_desc_ele.text
                    print(f"[{idx + 1}] 店铺介绍: {shop_description[:50]}...")  # 只显示前50个字符
                else:
                    print(f"[{idx + 1}] 未找到店铺介绍元素")
            else:
                print(f"[{idx + 1}] 未找到店铺按钮")
        except Exception as e:
            print(f"[{idx + 1}] 获取店铺介绍时出错: {e}")

        # 关闭当前标签页
        tab.close()
        print(f"[{idx + 1}] 标签页已关闭")

        if img_paths:
            print(f"[{idx + 1}] 准备处理 {len(img_paths)} 张图片")
            save_to_excel(
                img_paths=img_paths,
                cell_positions=cell_positions,
                output_filename=output_filename,
                product_url=url,
                row_number=idx + 2,
                status=status,
                run_key_data=run_key_row_data,
                run_key_titles=run_key_titles,
                green_price=green_price,
                gray_price=gray_price,
                shop_description=shop_description,
                price_html=price_html,
                run_key_path=run_key_path
            )
        else:
            print(f"[{idx + 1}] 未找到有效图片，仅保存链接和状态")
            save_to_excel(
                img_paths=[],
                cell_positions=[],
                output_filename=output_filename,
                product_url=url,
                row_number=idx + 2,
                status=status,
                run_key_data=run_key_row_data,
                run_key_titles=run_key_titles,
                green_price=green_price,
                gray_price=gray_price,
                shop_description=shop_description,
                price_html=price_html,
                run_key_path=run_key_path
            )

        time.sleep(0.5)  # 页面间等待

    except Exception as e:
        print(f"[{idx + 1}] 处理URL {url} 时发生错误: {e}")
        # 尝试关闭标签页（如果存在）
        if 'tab' in locals():
            try:
                tab.close()
            except:
                pass

        # 保存商品信息（无图片）
        save_to_excel(
            img_paths=[],
            cell_positions=[],
            output_filename=output_filename,
            product_url=url,
            row_number=idx + 2,
            status=f"错误: {str(e)[:50]}",  # 记录错误信息
            run_key_data=run_key_row_data,
            run_key_titles=run_key_titles,
            green_price=green_price,
            gray_price=gray_price,
            shop_description=shop_description,
            price_html=price_html,
            run_key_path=run_key_path
        )
    finally:
        # 无论成功或失败，更新处理状态
        update_processing_status(run_key_path, excel_row)


def update_processing_status(run_key_path, excel_row):
    """
    更新原始Excel文件中的处理状态

    参数:
    - run_key_path: 原始Excel文件路径
    - excel_row: 需要更新的行号（Excel中的实际行号）
    """
    # 使用锁保护Excel操作
    with excel_lock:
        try:
            # 加载工作簿
            wb = load_workbook(run_key_path)
            ws = wb.active

            # 查找"is_processed"列
            processed_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=6, column=col).value == "is_processed":
                    processed_col = col
                    break

            # 如果没找到处理状态列，创建它
            if processed_col is None:
                print("未找到处理状态列，创建新列")
                processed_col = ws.max_column + 1
                # 在标题行(第6行)添加列标题
                ws.cell(row=6, column=processed_col, value="is_processed")
                # 为之前的所有行添加默认值'n'
                for row_idx in range(7, ws.max_row + 1):
                    ws.cell(row=row_idx, column=processed_col, value='n')

            # 更新当前行的状态为'y'
            ws.cell(row=excel_row, column=processed_col, value='y')

            # 保存工作簿
            wb.save(run_key_path)
            print(f"已更新行 {excel_row} 的处理状态为 'y'")

        except Exception as e:
            print(f"更新处理状态时出错: {e}")
        finally:
            if 'wb' in locals():
                try:
                    wb.close()
                except:
                    pass


def get_processed_status(run_key_path):
    """
    获取原始Excel文件中的处理状态信息

    参数:
    - run_key_path: 原始Excel文件路径

    返回:
    - processed_col: 处理状态列的列号（Excel列号）
    - processed_rows: 已处理行的行号列表
    """
    processed_col = None
    processed_rows = []

    try:
        # 加载工作簿
        wb = load_workbook(run_key_path)
        ws = wb.active

        # 查找"is_processed"列
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=6, column=col).value == "is_processed":
                processed_col = col
                break

        # 如果找到处理状态列，获取所有已处理的行
        if processed_col:
            for row in range(7, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=processed_col).value
                if cell_value == 'y':
                    processed_rows.append(row)

        wb.close()

    except Exception as e:
        print(f"获取处理状态时出错: {e}")

    return processed_col, processed_rows


def prepare_run_key_data(run_key_path):
    """
    准备原始数据表，添加处理状态列（如果需要）

    参数:
    - run_key_path: 原始Excel文件路径

    返回:
    - df_run_key: 处理后的DataFrame
    - processed_rows: 已处理行的行号列表
    """
    # 获取处理状态信息
    processed_col, processed_rows = get_processed_status(run_key_path)

    # 读取数据部分
    df_run_key = pd.read_excel(run_key_path, header=5)

    # 如果没找到处理状态列，添加新列
    if processed_col is None:
        # 计算处理状态列的位置（最后一列之后）
        processed_col_idx = len(df_run_key.columns)

        # 在DataFrame中添加新列
        df_run_key["is_processed"] = 'n'

        # 使用openpyxl更新原始Excel文件
        try:
            wb = load_workbook(run_key_path)
            ws = wb.active

            # 添加列标题
            new_col = ws.max_column + 1
            ws.cell(row=6, column=new_col, value="is_processed")

            # 添加默认值
            for row in range(7, ws.max_row + 1):

                ws.cell(row=row, column=new_col, value='n')

            wb.save(run_key_path)
            wb.close()
            print("已添加处理状态列到原始Excel文件")

        except Exception as e:
            print(f"更新原始Excel文件时出错: {e}")

    return df_run_key, processed_rows


# 主程序
if __name__ == "__main__":
    # file_path = input("输入目录路径：")
    # xlsx_name = input("输入文件名称：")
    max_workers = int(input("输入最大线程数（推荐2-5）：") or 2)
    # 初始化输出文件
    output_filename = "结果.xlsx"

    # 确保目录存在
    # download_dir = file_path
    # if not os.path.exists(download_dir):
        # os.makedirs(download_dir)
        # print(f"创建目录: {download_dir}")

    # 创建图片目录
    image_dir = "图片"
    if not os.path.exists(image_dir):
        os.makedirs(image_dir)
        print(f"创建图片目录: {image_dir}")

    # 准备原始数据表
    run_key_path = "run_key.xlsx"
    if not os.path.exists(run_key_path):
        print(f"错误：找不到产品链接文件 {run_key_path}")
        exit(1)

    # 准备原始数据表，添加处理状态列（如果需要）
    df_run_key, processed_rows = prepare_run_key_data(run_key_path)
    print(f"原始数据表准备完成，已处理行数: {len(processed_rows)}")

    # 获取列标题
    run_key_titles = df_run_key.columns.tolist()

    # 查找产品链接列
    link_columns = [col for col in df_run_key.columns if "产品链接" in col]
    if not link_columns:
        link_columns = [col for col in df_run_key.columns if "链接" in col]

    if not link_columns:
        print("错误：找不到包含'产品链接'的列")
        exit(1)

    link_column = link_columns[0]
    print(f"使用 '{link_column}' 列作为产品链接列")

    # 获取产品链接列表
    product_links = df_run_key[link_column].tolist()
    print(f"找到 {len(product_links)} 个产品链接")

    # 用于登录
    page_dl = ChromiumPage()
    page_dl.get(r"https://www.ozon.ru/product")
    input("确认完成，按任意键然后回车：")
    page_dl.quit()
    time.sleep(2)

    # 创建主浏览器实例
    main_page = ChromiumPage()
    print("主浏览器实例已创建")


    try:
        # 使用线程池处理每个产品链接
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = []

            # 计算需要处理的行
            rows_to_process = []
            for idx, url in enumerate(product_links):
                excel_row = idx + 7  # Excel中的实际行号（标题在第6行，数据从第7行开始）

                # 跳过已处理的行
                if excel_row in processed_rows:
                    print(f"跳过已处理的行 {excel_row}")
                    continue

                # 获取run_key表对应行的数据
                run_key_row_data = df_run_key.iloc[idx].tolist() if idx < len(df_run_key) else None

                # 提交任务到线程池
                future = executor.submit(
                    process_single_url,
                    url=url,
                    idx=idx,
                    output_filename=output_filename,
                    image_dir=image_dir,
                    run_key_row_data=run_key_row_data,
                    run_key_titles=run_key_titles,
                    run_key_path=run_key_path,
                    max_workers=max_workers,
                    main_page=main_page,
                    excel_row=excel_row
                )
                futures.append(future)
                # 控制任务提交速度，避免瞬间提交过多任务
                time.sleep(0.5)

            # 等待所有任务完成
            for future in as_completed(futures):
                try:
                    future.result()  # 获取结果，如果有异常会在这里抛出
                except Exception as e:
                    print(f"任务执行出错: {e}")

    finally:
        # 确保关闭浏览器
        main_page.quit()
        print("主浏览器已关闭")
    print("所有任务处理完成！")